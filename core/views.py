from django.utils import timezone
from django.core.cache import cache
from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError as DRFValidationError
from django.db.models import Q, Count, F, Case, When, IntegerField
from django.db import transaction
from django.http import HttpResponse
from rest_framework.decorators import action
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from .models import Child, GalleryImage, Payment, Post, Attendance, DailyMenu, FacilityClosure, SpecialActivity, PostComment, GalleryItem, Group, RecurringPayment, Preschool
from .serializers import ChildSerializer, PaymentSerializer, RecurringPaymentSerializer, PostSerializer, AttendanceSerializer, FacilityClosureSerializer, SpecialActivitySerializer, DailyMenuSerializer, PostCommentSerializer, GalleryItemSerializer, GroupSerializer, PreschoolSerializer
from .group_deletion import delete_group_with_related_data
from users.permissions import IsDirector, IsDirectorOrTeacher
from users.models import User, EmailNotificationEventType
from users.email_notifications import queue_parent_email_notification
from rest_framework.views import APIView
from communication.models import Message
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from .meal_payments import generate_current_month_meal_payments


def broadcast_notification_summary_changed(user_ids=None):
    channel_layer = get_channel_layer()
    if not channel_layer:
        return

    if user_ids is None:
        user_ids = User.objects.values_list('id', flat=True)

    for user_id in user_ids:
        async_to_sync(channel_layer.group_send)(
            f'user_{int(user_id)}',
            {
                'type': 'chat.notification_summary_changed',
            }
        )


def increment_schedule_change_notification(user_ids):
    for user_id in set(user_ids):
        cache_key = f'notification_schedule_extra_{int(user_id)}'
        current_value = int(cache.get(cache_key, 0) or 0)
        cache.set(cache_key, current_value + 1, timeout=60 * 60 * 24 * 30)

class ChildViewSet(viewsets.ModelViewSet):
    serializer_class = ChildSerializer
    permission_classes = [permissions.IsAuthenticated]

    # USUNĘLIŚMY LINIĘ: http_method_names = [...] 
    # Teraz domyślnie dozwolone jest wszystko, ale ograniczymy to poniżej.

    def get_queryset(self):
        user = self.request.user
        if user.is_director:
            return Child.objects.all()
        return user.child.all() # Upewnij się, że masz tu .children.all() (zależnie od related_name w models.py, chyba zmienialiśmy na .child.all()?)
        # SPRAWDŹ models.py: 
        # Jeśli w models.py Child ma: parents = ManyToManyField(..., related_name='children') -> to użyj user.children.all()
        # Jeśli w models.py Child ma: parents = ManyToManyField(..., related_name='child') -> to użyj user.child.all()
        # (Wcześniej poprawialiśmy błąd na .child.all(), więc trzymajmy się tego co działa u Ciebie)

    def get_permissions(self):
        """
        Dyrektor: Pełen dostęp (Create, Delete, Update).
        Rodzic: Tylko odczyt (Get) i aktualizacja medyczna (Patch).
        """
        if self.action in ['create', 'destroy']:
            return [IsDirector()] # Tylko dyrektor może tworzyć/usuwać
        return super().get_permissions()

    def update(self, request, *args, **kwargs):
        # Logika dla Rodzica (zabezpieczenie pól)
        if not request.user.is_director:
            # Pozwalamy edytować TYLKO medical_info
            allowed_data = {'medical_info': request.data.get('medical_info', request.data.get('medical_info', ''))}
            
            # Jeśli rodzic próbuje zmienić coś innego, ignorujemy to
            # (Nadpisujemy dane wejściowe tylko dozwolonym polem)
            serializer = self.get_serializer(self.get_object(), data=allowed_data, partial=True)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data)
            
        # Logika dla Dyrektora (pełna edycja)
        return super().update(request, *args, **kwargs)

class PaymentViewSet(viewsets.ModelViewSet):
    serializer_class = PaymentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        child_id = self.request.query_params.get('child_id')

        if user.is_director:
            queryset = Payment.objects.all()
            if child_id:
                queryset = queryset.filter(child_id=child_id)
            return queryset

        queryset = Payment.objects.filter(child__parents=user)
        if child_id:
            queryset = queryset.filter(child_id=child_id)
        return queryset
        
    def perform_update(self, serializer):
        # Zabezpieczenie: tylko dyrektor może zmienić status i datę opłacenia
        if not self.request.user.is_director:
            if 'is_paid' in serializer.validated_data:
                serializer.validated_data.pop('is_paid')
            if 'payment_date' in serializer.validated_data:
                serializer.validated_data.pop('payment_date')
        serializer.save()

    def perform_create(self, serializer):
        if not self.request.user.is_director:
            serializer.validated_data.pop('is_paid', None)
            serializer.validated_data.pop('payment_date', None)

        payment = serializer.save()

        parent_ids = payment.child.parents.values_list('id', flat=True)
        director_ids = User.objects.filter(is_director=True).values_list('id', flat=True)
        target_ids = set(parent_ids) | set(director_ids)
        broadcast_notification_summary_changed(target_ids)

        for parent in payment.child.parents.all():
            queue_parent_email_notification(parent, EmailNotificationEventType.PAYMENTS)

    @action(detail=False, methods=['get'], url_path='generate-title')
    def generate_title(self, request):
        child_id = request.query_params.get('child_id')
        if not child_id:
            return Response({'detail': 'Parametr child_id jest wymagany.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            child_id_int = int(child_id)
        except (TypeError, ValueError):
            return Response({'detail': 'Parametr child_id musi być liczbą całkowitą.'}, status=status.HTTP_400_BAD_REQUEST)

        child_queryset = Child.objects.filter(id=child_id_int)
        if not request.user.is_director:
            child_queryset = child_queryset.filter(parents=request.user)

        child = child_queryset.first()
        if not child:
            return Response({'detail': 'Nie znaleziono dziecka dla podanego identyfikatora.'}, status=status.HTTP_404_NOT_FOUND)

        generated_title = Payment(
            child=child,
            amount=Decimal('0.00'),
            description='Podgląd tytułu',
        ).generate_unique_title()

        return Response({'payment_title': generated_title}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='generate-meal-payments', permission_classes=[IsDirector])
    def generate_meal_payments(self, request):
        raw_adjustments = request.data.get('adjustments', [])

        try:
            adjustments_by_child = self._parse_manual_adjustments(raw_adjustments)
        except DRFValidationError as exc:
            detail = exc.detail
            if isinstance(detail, dict):
                detail = detail.get('detail', detail)
            return Response({'detail': str(detail)}, status=status.HTTP_400_BAD_REQUEST)

        today = timezone.now().date()
        try:
            with transaction.atomic():
                result = generate_current_month_meal_payments(
                    today=today,
                    include_previous_month_absences=True,
                )
                adjusted_count = self._apply_manual_adjustments(
                    meal_period=result['meal_period'],
                    adjustments_by_child=adjustments_by_child,
                )
        except DRFValidationError as exc:
            detail = exc.detail
            if isinstance(detail, dict):
                detail = detail.get('detail', detail)
            return Response({'detail': str(detail)}, status=status.HTTP_400_BAD_REQUEST)

        detail_message = (
            f"Wygenerowano {result['created_count']} płatności za posiłki. "
            f"Pominięto {result['skipped_count']} istniejących."
        )
        if adjusted_count:
            detail_message += f" Zastosowano {adjusted_count} korekt ręcznych."

        return Response(
            {
                'detail': detail_message,
                'meal_period': result['meal_period'].isoformat(),
                'created_count': result['created_count'],
                'skipped_count': result['skipped_count'],
                'adjusted_count': adjusted_count,
            },
            status=status.HTTP_200_OK,
        )

    def _parse_manual_adjustments(self, raw_adjustments):
        if raw_adjustments in (None, ''):
            return {}

        if not isinstance(raw_adjustments, list):
            raise DRFValidationError({'detail': 'Pole adjustments musi być listą.'})

        parsed = {}
        for index, item in enumerate(raw_adjustments, start=1):
            if not isinstance(item, dict):
                raise DRFValidationError({'detail': f'Korekta #{index} ma nieprawidłowy format.'})

            child_id = item.get('child_id')
            try:
                child_id = int(child_id)
            except (TypeError, ValueError):
                raise DRFValidationError({'detail': f'Korekta #{index}: child_id musi być liczbą całkowitą.'})

            if child_id in parsed:
                raise DRFValidationError({'detail': f'Korekta dla dziecka ID={child_id} została podana więcej niż raz.'})

            raw_delta = item.get('amount_delta')
            try:
                amount_delta = Decimal(str(raw_delta)).quantize(Decimal('0.01'))
            except (InvalidOperation, TypeError, ValueError):
                raise DRFValidationError({'detail': f'Korekta #{index}: amount_delta ma nieprawidłowy format.'})

            if amount_delta == Decimal('0.00'):
                raise DRFValidationError({'detail': f'Korekta #{index}: amount_delta nie może być równe 0.'})

            reason = (item.get('reason') or '').strip()
            if not reason:
                raise DRFValidationError({'detail': f'Korekta #{index}: podaj powód korekty.'})

            parsed[child_id] = {
                'amount_delta': amount_delta,
                'reason': reason,
            }

        if not parsed:
            return {}

        allowed_children = Child.objects.filter(
            id__in=list(parsed.keys()),
            uses_meals=True,
        )
        allowed_by_id = {child.id: child for child in allowed_children}
        missing_ids = [child_id for child_id in parsed.keys() if child_id not in allowed_by_id]
        if missing_ids:
            raise DRFValidationError(
                {'detail': f'Nie można dodać korekty. Dzieci niedostępne dla wyżywienia: {", ".join(map(str, missing_ids))}.'}
            )

        for child_id in parsed.keys():
            parsed[child_id]['child_name'] = f"{allowed_by_id[child_id].first_name} {allowed_by_id[child_id].last_name}".strip()

        return parsed

    def _apply_manual_adjustments(self, meal_period, adjustments_by_child):
        if not adjustments_by_child:
            return 0

        adjusted_count = 0
        for child_id, adjustment in adjustments_by_child.items():
            payment = Payment.objects.select_for_update().filter(
                child_id=child_id,
                meal_period=meal_period,
            ).first()

            if not payment:
                raise DRFValidationError(
                    {'detail': f'Brak płatności wyżywieniowej do korekty dla dziecka: {adjustment["child_name"]}.'}
                )

            updated_amount = (payment.amount + adjustment['amount_delta']).quantize(Decimal('0.01'))
            if updated_amount < Decimal('0.00'):
                raise DRFValidationError(
                    {'detail': f'Korekta dla dziecka {adjustment["child_name"]} obniża kwotę poniżej 0.00 zł.'}
                )

            payment.amount = updated_amount
            correction_note = (
                f" Korekta dyrektora: {adjustment['amount_delta']:+.2f} zł "
                f"(powód: {adjustment['reason']})."
            )
            if correction_note not in payment.description:
                payment.description = f"{payment.description}{correction_note}".strip()

            payment.save(update_fields=['amount', 'description'])
            adjusted_count += 1

        return adjusted_count


class RecurringPaymentViewSet(viewsets.ModelViewSet):
    serializer_class = RecurringPaymentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.is_director:
            return RecurringPayment.objects.all()
        return RecurringPayment.objects.filter(children__parents=user).distinct()

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsDirector()]
        return super().get_permissions()

class PostViewSet(viewsets.ModelViewSet): # <--- ZMIANA 1: ModelViewSet (zamiast ReadOnly)
    """
    Zwraca listę postów (tablicę).
    Dyrektor: Pełny dostęp (CRUD).
    Rodzic: Widzi posty ogólne ORAZ przypisane do grup jego dzieci. Może lajkować/komentować.
    """
    serializer_class = PostSerializer
    # Domyślnie wymagamy zalogowania (dla listowania, lajków itp.)
    permission_classes = [permissions.IsAuthenticated]

    # --- ZMIANA 2: OCHRONA ZAPISU ---
    def get_permissions(self):
        """
        Dynamiczne przydzielanie uprawnień:
        - Edycja/Usuwanie/Tworzenie -> Dyrektor lub Nauczyciel.
        - Czytanie/Lajkowanie/Komentowanie -> Każdy zalogowany.
        """
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsDirectorOrTeacher()]
        return super().get_permissions()

    # --- TWOJA ORYGINALNA LOGIKA FILTROWANIA (BEZ ZMIAN) ---
    def get_queryset(self):
        user = self.request.user
        child_id = self.request.query_params.get('child_id')
        
        # 1. Jeśli to Dyrektor -> widzi wszystko
        if user.is_director or user.is_teacher:
            queryset = Post.objects.all()
            if child_id:
                try:
                    child = Child.objects.get(id=child_id)
                    return queryset.filter(Q(target_group__isnull=True) | Q(target_group=child.group)).distinct()
                except Child.DoesNotExist:
                    return queryset.none()
            return queryset
        
        # 2. Jeśli to Rodzic -> pobieramy wszystkie jego dzieci
        children = user.child.all()
        
        # Jeśli rodzic nie ma przypisanych dzieci, widzi tylko posty ogólne
        if not children.exists():
            return Post.objects.filter(target_group__isnull=True)

        if child_id:
            selected_child = children.filter(id=child_id).first()
            if not selected_child:
                return Post.objects.none()
            return Post.objects.filter(
                Q(target_group__isnull=True) | Q(target_group=selected_child.group)
            ).distinct()

        # 3. Zbieramy grupy wszystkich dzieci rodzica do jednej listy
        parent_groups = [child.group for child in children]
        
        # 4. Filtrujemy posty
        return Post.objects.filter(
            Q(target_group__isnull=True) | Q(target_group__in=parent_groups)
        ).distinct()

    # --- TWOJE ORYGINALNE AKCJE (BEZ ZMIAN) ---

    def _get_post_parent_targets(self, post):
        if post.target_group_id:
            return User.objects.filter(is_parent=True, child__group_id=post.target_group_id).distinct()
        return User.objects.filter(is_parent=True).distinct()

    def perform_create(self, serializer):
        post = serializer.save()
        for parent in self._get_post_parent_targets(post):
            queue_parent_email_notification(parent, EmailNotificationEventType.POSTS)

    @action(detail=True, methods=['post'])
    def like(self, request, pk=None):
        post = self.get_object()
        user = request.user

        if post.likes.filter(id=user.id).exists():
            post.likes.remove(user)
            liked = False
        else:
            post.likes.add(user)
            liked = True

        return Response({
            'liked': liked, 
            'likes_count': post.likes.count()
        })

    @action(detail=True, methods=['post'])
    def comment(self, request, pk=None):
        post = self.get_object()
        content = request.data.get('content')

        if not content:
            return Response({'error': 'Treść komentarza jest wymagana'}, status=status.HTTP_400_BAD_REQUEST)

        comment = PostComment.objects.create(
            post=post,
            author=request.user,
            content=content
        )
        
        serializer = PostCommentSerializer(comment)
        return Response(serializer.data, status=status.HTTP_201_CREATED)    

    def update(self, request, *args, **kwargs):
        instance = self.get_object()

        instance.title = request.data.get('title', instance.title)
        instance.content = request.data.get('content', instance.content)

        target_group_id = request.data.get('target_group')
        if target_group_id:
            instance.target_group_id = target_group_id
        else:
            instance.target_group = None

        new_image = request.FILES.get('image')
        if new_image:
            if instance.image:
                instance.image.delete(save=False)
            instance.image = new_image

        delete_image_value = str(request.data.get('delete_image', '')).strip().lower()
        delete_image = delete_image_value in ['1', 'true', 'yes', 'on']

        if delete_image and not new_image:
            if instance.image:
                instance.image.delete(save=False)
            instance.image = None

        instance.save()

        for parent in self._get_post_parent_targets(instance):
            queue_parent_email_notification(parent, EmailNotificationEventType.POSTS)

        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    def perform_destroy(self, instance):
        instance.delete()
      
class AttendanceViewSet(viewsets.ModelViewSet):
    serializer_class = AttendanceSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        child_id = self.request.query_params.get('child_id')
        
        # Dyrektor widzi listę dla całego przedszkola
        if user.is_director:
            queryset = Attendance.objects.all()
            if child_id:
                queryset = queryset.filter(child_id=child_id)
            return queryset
            
        queryset = Attendance.objects.filter(child__parents=user)
        if child_id:
            queryset = queryset.filter(child_id=child_id)
        return queryset

    @action(detail=False, methods=['get'], url_path='meal-report', permission_classes=[IsDirector])
    def meal_report(self, request):
        today = timezone.now().date()

        report_month_last_day = today.replace(day=1) - timedelta(days=1)
        report_month_first_day = report_month_last_day.replace(day=1)

        billing_month_first_day = today.replace(day=1)
        billing_month_last_day = (billing_month_first_day + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        billing_month_locative = {
            1: 'styczniu',
            2: 'lutym',
            3: 'marcu',
            4: 'kwietniu',
            5: 'maju',
            6: 'czerwcu',
            7: 'lipcu',
            8: 'sierpniu',
            9: 'wrzesniu',
            10: 'pazdzierniku',
            11: 'listopadzie',
            12: 'grudniu',
        }

        closures_in_report_month = set(
            FacilityClosure.objects.filter(
                date__range=[report_month_first_day, report_month_last_day]
            ).values_list('date', flat=True)
        )

        closures_in_billing_month = set(
            FacilityClosure.objects.filter(
                date__range=[billing_month_first_day, billing_month_last_day]
            ).values_list('date', flat=True)
        )

        billing_business_days = 0
        day_cursor = billing_month_first_day
        while day_cursor <= billing_month_last_day:
            if day_cursor.isoweekday() <= 5 and day_cursor not in closures_in_billing_month:
                billing_business_days += 1
            day_cursor += timedelta(days=1)

        groups = Group.objects.all().order_by('name')
        children_by_group = {
            group.id: list(
                group.children.filter(uses_meals=True)
                .order_by('last_name', 'first_name', 'id')
            )
            for group in groups
        }

        report_days = []
        day_cursor = report_month_first_day
        while day_cursor <= report_month_last_day:
            report_days.append(day_cursor)
            day_cursor += timedelta(days=1)

        attendance_rows = Attendance.objects.filter(
            date__range=[report_month_first_day, report_month_last_day],
            status='absent',
            child__uses_meals=True,
        ).values('child_id', 'date')

        absences_by_child = {}
        for row in attendance_rows:
            absences_by_child.setdefault(row['child_id'], set()).add(row['date'])

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Raport wyzywienia'

        title = (
            f'Raport nieobecnosci: {report_month_first_day.strftime("%m/%Y")}; '
            f'naliczenie: {billing_month_first_day.strftime("%m/%Y")}'
        )
        max_col = len(report_days) + 4
        sheet.cell(row=1, column=1, value=title)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        sheet.cell(row=1, column=1).font = Font(bold=True)

        current_row = 3
        preschool_total = Decimal('0.00')
        abs_col = len(report_days) + 2
        rate_col = len(report_days) + 3
        due_col = len(report_days) + 4

        for group in groups:
            children = children_by_group[group.id]

            sheet.cell(row=current_row, column=1, value=group.name)
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_col)
            sheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            sheet.cell(row=current_row, column=1, value='data').font = Font(bold=True)
            for idx, month_day in enumerate(report_days, start=2):
                sheet.cell(row=current_row, column=idx, value=month_day.day).alignment = Alignment(horizontal='center')

            sheet.cell(row=current_row, column=abs_col, value='nieobecnosci').font = Font(bold=True)
            sheet.cell(row=current_row, column=rate_col, value='stawka wyzywienia').font = Font(bold=True)
            sheet.cell(row=current_row, column=due_col, value='do zaplaty').font = Font(bold=True)
            current_row += 1

            group_total = Decimal('0.00')

            for child in children:
                child_absence_dates = absences_by_child.get(child.id, set())
                billable_absences = sum(
                    1
                    for absence_date in child_absence_dates
                    if absence_date.isoweekday() <= 5 and absence_date not in closures_in_report_month
                )

                billable_days = max(billing_business_days - billable_absences, 0)
                amount_due = (Decimal(billable_days) * child.meal_rate).quantize(Decimal('0.01'))
                group_total += amount_due

                sheet.cell(
                    row=current_row,
                    column=1,
                    value=f'{child.first_name} {child.last_name}',
                )

                for idx, month_day in enumerate(report_days, start=2):
                    mark = 'N' if month_day in child_absence_dates else ''
                    sheet.cell(row=current_row, column=idx, value=mark).alignment = Alignment(horizontal='center')

                sheet.cell(row=current_row, column=abs_col, value=billable_absences)
                sheet.cell(row=current_row, column=rate_col, value=float(child.meal_rate))
                sheet.cell(row=current_row, column=due_col, value=float(amount_due))
                current_row += 1

            sheet.cell(row=current_row, column=rate_col, value='calkowita suma w grupie').font = Font(bold=True)
            sheet.cell(row=current_row, column=due_col, value=float(group_total)).font = Font(bold=True)
            preschool_total += group_total
            current_row += 2

        sheet.cell(row=current_row, column=rate_col, value='calkowita suma w przedszkolu').font = Font(bold=True)
        sheet.cell(row=current_row, column=due_col, value=float(preschool_total)).font = Font(bold=True)

        current_row += 2
        sheet.cell(row=current_row, column=1, value='ILOSC DNI').font = Font(bold=True)
        sheet.cell(row=current_row, column=2, value=billing_business_days)
        sheet.cell(
            row=current_row,
            column=3,
            value=(
                f'dni robocze w {billing_month_locative[billing_month_first_day.month]} '
                f'(bez weekendow i dni wolnych)'
            )
        )

        for col_idx in range(1, max_col + 1):
            if col_idx == 1:
                width = 28
            elif col_idx <= len(report_days) + 1:
                width = 4
            else:
                width = 20
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        money_format = '#,##0.00'
        for row in range(4, current_row + 1):
            due_value = sheet.cell(row=row, column=due_col).value
            rate_value = sheet.cell(row=row, column=rate_col).value
            if isinstance(due_value, (int, float)):
                sheet.cell(row=row, column=due_col).number_format = money_format
            if isinstance(rate_value, (int, float)):
                sheet.cell(row=row, column=rate_col).number_format = money_format

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = (
            f'raport_nieobecnosci_{report_month_first_day.strftime("%Y_%m")}'
            f'_naliczenie_{billing_month_first_day.strftime("%Y_%m")}.xlsx'
        )
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
class FacilityClosureViewSet(viewsets.ModelViewSet):
    """
    Zwraca listę dni, kiedy przedszkole jest zamknięte.
    """
    queryset = FacilityClosure.objects.all()
    serializer_class = FacilityClosureSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_permissions(self):
        # Dyrektor i nauczyciel mogą edytować
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsDirectorOrTeacher()]
        return super().get_permissions()

    def perform_create(self, serializer):
        serializer.save()
        broadcast_notification_summary_changed()

        for parent in User.objects.filter(is_parent=True).exclude(email='').exclude(email__isnull=True):
            queue_parent_email_notification(parent, EmailNotificationEventType.CALENDAR)

    def perform_update(self, serializer):
        serializer.save()
        broadcast_notification_summary_changed()

        for parent in User.objects.filter(is_parent=True).exclude(email='').exclude(email__isnull=True):
            queue_parent_email_notification(parent, EmailNotificationEventType.CALENDAR)

    def perform_destroy(self, instance):
        instance.delete()
        broadcast_notification_summary_changed()

class SpecialActivityViewSet(viewsets.ModelViewSet):
    """
    Zwraca zajęcia dodatkowe.
    Rodzic widzi zajęcia przypisane do grup jego dzieci.
    Dyrektor widzi wszystko.
    """
    serializer_class = SpecialActivitySerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_permissions(self):
        # Dyrektor i nauczyciel mogą edytować
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsDirectorOrTeacher()]
        return super().get_permissions()

    def get_queryset(self):
        user = self.request.user
        child_id = self.request.query_params.get('child_id')
        
        # Dyrektor widzi cały kalendarz
        if user.is_director or user.is_teacher:
            queryset = SpecialActivity.objects.all()
            if child_id:
                try:
                    child = Child.objects.get(id=child_id)
                    return queryset.filter(groups=child.group).distinct()
                except Child.DoesNotExist:
                    return queryset.none()
            return queryset
        
        # Rodzic: pobieramy grupy jego dzieci
        children = user.child.all()
        if not children.exists():
            return SpecialActivity.objects.none()

        if child_id:
            selected_child = children.filter(id=child_id).first()
            if not selected_child:
                return SpecialActivity.objects.none()
            return SpecialActivity.objects.filter(groups=selected_child.group).distinct()
            
        parent_groups = [child.group for child in children]
        
        # Filtrujemy zajęcia, które są przypisane do którejkolwiek z tych grup
        # distinct() jest ważne przy ManyToMany, żeby nie dublować wyników
        return SpecialActivity.objects.filter(groups__in=parent_groups).distinct()

    def _get_activity_notification_target_ids(self, group_ids):
        normalized_group_ids = set(group_ids)
        if not normalized_group_ids:
            normalized_group_ids = set(Group.objects.values_list('id', flat=True))

        parent_ids = User.objects.filter(child__group_id__in=normalized_group_ids).values_list('id', flat=True).distinct()
        director_ids = User.objects.filter(is_director=True).values_list('id', flat=True)
        return set(parent_ids) | set(director_ids)

    def perform_create(self, serializer):
        activity = serializer.save()

        group_ids = activity.groups.values_list('id', flat=True)
        target_ids = self._get_activity_notification_target_ids(group_ids)
        broadcast_notification_summary_changed(target_ids)

        parent_targets = User.objects.filter(id__in=target_ids, is_parent=True)
        for parent in parent_targets:
            queue_parent_email_notification(parent, EmailNotificationEventType.SCHEDULE)

    def perform_update(self, serializer):
        previous_group_ids = serializer.instance.groups.values_list('id', flat=True)
        activity = serializer.save()
        updated_group_ids = activity.groups.values_list('id', flat=True)

        all_relevant_group_ids = set(previous_group_ids) | set(updated_group_ids)
        target_ids = self._get_activity_notification_target_ids(all_relevant_group_ids)
        increment_schedule_change_notification(target_ids)
        broadcast_notification_summary_changed(target_ids)

        parent_targets = User.objects.filter(id__in=target_ids, is_parent=True)
        for parent in parent_targets:
            queue_parent_email_notification(parent, EmailNotificationEventType.SCHEDULE)

    def perform_destroy(self, instance):
        instance.delete()
        broadcast_notification_summary_changed()
    
class DailyMenuViewSet(viewsets.ModelViewSet):
    """
    Zwraca jadłospis.
    Można filtrować po dacie, np. ?date__gte=2025-11-01&date__lte=2025-11-07
    """
    queryset = DailyMenu.objects.all().order_by('-week_start_date')
    serializer_class = DailyMenuSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_permissions(self):
        # Tylko dyrektor może edytować
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsDirector()]
        return super().get_permissions()
    
    # Filtrowanie po zakresie dat (zwraca jadłospisy, które nachodzą na podany zakres)
    def get_queryset(self):
        queryset = super().get_queryset()
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        
        if start_date and end_date:
            try:
                range_start = date.fromisoformat(start_date)
                range_end = date.fromisoformat(end_date)
            except ValueError:
                return queryset.none()

            latest_possible_week_start = range_end
            earliest_possible_week_start = range_start - timedelta(days=4)
            return queryset.filter(
                week_start_date__lte=latest_possible_week_start,
                week_start_date__gte=earliest_possible_week_start,
            )
        return queryset
    
class GalleryViewSet(viewsets.ModelViewSet):
    """
    Zarządzanie albumami (Galeria).
    Dyrektor: Pełny dostęp (CRUD + obsługa plików).
    Rodzic: Tylko odczyt + lajki.
    """
    serializer_class = GalleryItemSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_permissions(self):
        # Dyrektor i nauczyciel mogą tworzyć/edytować/usuwać albumy
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsDirectorOrTeacher()]
        return super().get_permissions()

    def get_queryset(self):
        user = self.request.user
        child_id = self.request.query_params.get('child_id')
        
        if user.is_director or user.is_teacher:
            queryset = GalleryItem.objects.all()
            if child_id:
                try:
                    child = Child.objects.get(id=child_id)
                    return queryset.filter(
                        Q(target_group__isnull=True) | Q(target_group=child.group)
                    ).distinct()
                except Child.DoesNotExist:
                    return queryset.none()
            return queryset
        
        children = user.child.all()
        if not children.exists():
            return GalleryItem.objects.filter(target_group__isnull=True)

        if child_id:
            selected_child = children.filter(id=child_id).first()
            if not selected_child:
                return GalleryItem.objects.none()
            return GalleryItem.objects.filter(
                Q(target_group__isnull=True) | Q(target_group=selected_child.group)
            ).distinct()

        parent_groups = [child.group for child in children]
        
        return GalleryItem.objects.filter(
            Q(target_group__isnull=True) | Q(target_group__in=parent_groups)
        ).distinct()

    def _get_gallery_parent_targets(self, album):
        if album.target_group_id:
            return User.objects.filter(is_parent=True, child__group_id=album.target_group_id).distinct()
        return User.objects.filter(is_parent=True).distinct()

    # --- AKCJA LAJKOWANIA ALBUMU ---
    @action(detail=True, methods=['post'])
    def like(self, request, pk=None):
        album = self.get_object()
        user = request.user

        if album.likes.filter(id=user.id).exists():
            album.likes.remove(user)
            liked = False
        else:
            album.likes.add(user)
            liked = True

        return Response({
            'liked': liked, 
            'likes_count': album.likes.count()
        })

    # --- NOWA METODA CREATE (dla wielu zdjęć z Frontendu) ---
    def create(self, request, *args, **kwargs):
        title = request.data.get('title')
        description = request.data.get('description', '')
        target_group_id = request.data.get('target_group')
        
        if not title:
            return Response({'title': 'Tytuł jest wymagany.'}, status=status.HTTP_400_BAD_REQUEST)
        
        album = GalleryItem.objects.create(
            title=title,
            description=description,
            target_group_id=target_group_id if target_group_id else None
        )
        
        # Pobieramy listę plików (zdjęć)
        images = request.FILES.getlist('images')
        
        # W pętli tworzymy obiekty GalleryImage
        for image_file in images:
            GalleryImage.objects.create(gallery_item=album, image=image_file)

        if album.target_group_id:
            parent_ids = User.objects.filter(child__group_id=album.target_group_id).values_list('id', flat=True).distinct()
            director_ids = User.objects.filter(is_director=True).values_list('id', flat=True)
            target_ids = set(parent_ids) | set(director_ids)
            broadcast_notification_summary_changed(target_ids)
        else:
            broadcast_notification_summary_changed()

        for parent in self._get_gallery_parent_targets(album):
            queue_parent_email_notification(parent, EmailNotificationEventType.GALLERY)
            
        serializer = self.get_serializer(album)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    # --- NOWA METODA UPDATE (dla edycji zdjęć) ---
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        
        # 1. Aktualizuj dane tekstowe
        instance.title = request.data.get('title', instance.title)
        instance.description = request.data.get('description', instance.description)
        
        target_group_id = request.data.get('target_group')
        if target_group_id:
            instance.target_group_id = target_group_id
        else:
            instance.target_group = None
        instance.save()
        
        # 2. Dodawanie nowych zdjęć
        new_images = request.FILES.getlist('images')
        for image_file in new_images:
            GalleryImage.objects.create(gallery_item=instance, image=image_file)

        # 3. Usuwanie starych zdjęć
        # Frontend wyśle listę ID zdjęć do usunięcia, np. 'deleted_images': [1, 5, 12]
        deleted_images_ids = request.data.getlist('deleted_images', [])
        if deleted_images_ids:
            GalleryImage.objects.filter(id__in=deleted_images_ids, gallery_item=instance).delete()

        if instance.target_group_id:
            parent_ids = User.objects.filter(child__group_id=instance.target_group_id).values_list('id', flat=True).distinct()
            director_ids = User.objects.filter(is_director=True).values_list('id', flat=True)
            target_ids = set(parent_ids) | set(director_ids)
            broadcast_notification_summary_changed(target_ids)
        else:
            broadcast_notification_summary_changed()

        for parent in self._get_gallery_parent_targets(instance):
            queue_parent_email_notification(parent, EmailNotificationEventType.GALLERY)
            
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    def perform_destroy(self, instance):
        if instance.target_group_id:
            parent_ids = User.objects.filter(child__group_id=instance.target_group_id).values_list('id', flat=True).distinct()
            director_ids = User.objects.filter(is_director=True).values_list('id', flat=True)
            target_ids = set(parent_ids) | set(director_ids)
        else:
            target_ids = None

        instance.delete()

        if target_ids is None:
            broadcast_notification_summary_changed()
        else:
            broadcast_notification_summary_changed(target_ids)
    
class CommentViewSet(viewsets.GenericViewSet):
    queryset = PostComment.objects.all()
    permission_classes = [permissions.IsAuthenticated]

    # AKCJA: Polub komentarz
    # POST /api/comments/{id}/like/
    @action(detail=True, methods=['post'])
    def like(self, request, pk=None):
        comment = self.get_object()
        user = request.user

        if comment.likes.filter(id=user.id).exists():
            comment.likes.remove(user)
            liked = False
        else:
            comment.likes.add(user)
            liked = True

        return Response({
            'liked': liked, 
            'likes_count': comment.likes.count()
        })

class GroupViewSet(viewsets.ModelViewSet): # Zmieniamy na ModelViewSet (pełny dostęp)
    serializer_class = GroupSerializer # Zakładam, że masz ten serializer w core/serializers.py
    # Domyślne uprawnienie: Zalogowany (żeby rodzic widział grupy)
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Group.objects.all()

    def get_permissions(self):
        """
        Dostosowujemy uprawnienia w zależności od akcji.
        - Przeglądanie (list, retrieve): Każdy zalogowany (Rodzic/Dyrektor)
        - Edycja/Usuwanie/Tworzenie: Tylko Dyrektor
        """
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsDirector()]
        return super().get_permissions()

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()

        if instance.children.exists():
            director_password = (request.data.get('director_password') or '').strip()
            if not director_password:
                return Response(
                    {'detail': 'Aby usunąć grupę z dziećmi, wpisz hasło dyrektora.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if not request.user.check_password(director_password):
                return Response(
                    {'detail': 'Nieprawidłowe hasło dyrektora.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        delete_group_with_related_data(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)
    
class DirectorStatsView(APIView):
    """
    Zwraca statystyki dla pulpitu dyrektora.
    """
    permission_classes = [IsDirector] # Tylko dyrektor

    def _build_attendance_series(self, days, group_id=None):
        today = timezone.localdate()
        start_date = today - timedelta(days=days - 1)

        children_qs = Child.objects.all()
        if group_id is not None:
            children_qs = children_qs.filter(group_id=group_id)

        total_children = children_qs.count()

        absence_qs = Attendance.objects.filter(
            status='absent',
            date__range=(start_date, today),
        )
        if group_id is not None:
            absence_qs = absence_qs.filter(child__group_id=group_id)

        absences_by_date = {
            entry['date']: entry['count']
            for entry in absence_qs.values('date').annotate(count=Count('id'))
        }

        series = []
        for day_index in range(days):
            current_date = start_date + timedelta(days=day_index)
            absent_count = int(absences_by_date.get(current_date, 0) or 0)
            present_count = max(total_children - absent_count, 0)
            attendance_rate = round((present_count / total_children) * 100, 1) if total_children else 0.0

            series.append({
                'date': current_date.isoformat(),
                'label': current_date.strftime('%d.%m'),
                'present': present_count,
                'absent': absent_count,
                'total': total_children,
                'attendance_rate': attendance_rate,
            })

        return series

    def _build_debt_stats(self):
        unpaid_payments = Payment.objects.filter(is_paid=False).select_related(
            'child',
            'child__group',
        ).prefetch_related('child__parents')

        total_outstanding = Decimal('0.00')
        group_totals = {}
        parent_map = {}

        for payment in unpaid_payments:
            payment_amount = payment.amount or Decimal('0.00')
            total_outstanding += payment_amount

            group_obj = payment.child.group
            group_key = int(group_obj.id)
            if group_key not in group_totals:
                group_totals[group_key] = {
                    'group_id': group_key,
                    'group_name': group_obj.name,
                    'amount': Decimal('0.00'),
                    'unpaid_items': 0,
                }
            group_totals[group_key]['amount'] += payment_amount
            group_totals[group_key]['unpaid_items'] += 1

            parent_users = list(payment.child.parents.all())
            for parent in parent_users:
                parent_key = int(parent.id)
                full_name = f"{parent.first_name} {parent.last_name}".strip() or parent.username
                if parent_key not in parent_map:
                    parent_map[parent_key] = {
                        'parent_id': parent_key,
                        'parent_name': full_name,
                        'amount': Decimal('0.00'),
                        'unpaid_items': 0,
                        'group_names': set(),
                        'debts': [],
                    }

                parent_map[parent_key]['amount'] += payment_amount
                parent_map[parent_key]['unpaid_items'] += 1
                parent_map[parent_key]['group_names'].add(group_obj.name)
                parent_map[parent_key]['debts'].append({
                    'payment_id': payment.id,
                    'payment_title': payment.payment_title,
                    'description': payment.description,
                    'amount': float(payment_amount),
                    'group_id': group_key,
                    'group_name': group_obj.name,
                    'child_id': payment.child_id,
                    'child_name': f"{payment.child.first_name} {payment.child.last_name}",
                    'created_at': payment.created_at.isoformat(),
                })

        debtors = []
        for entry in parent_map.values():
            entry['amount'] = float(entry['amount'])
            entry['group_names'] = sorted(list(entry['group_names']))
            entry['debts'].sort(key=lambda debt: debt['created_at'])
            debtors.append(entry)

        debtors.sort(key=lambda debtor: debtor['amount'], reverse=True)

        by_group = []
        for group_entry in group_totals.values():
            by_group.append({
                'group_id': group_entry['group_id'],
                'group_name': group_entry['group_name'],
                'amount': float(group_entry['amount']),
                'unpaid_items': group_entry['unpaid_items'],
            })
        by_group.sort(key=lambda group_entry: group_entry['amount'], reverse=True)

        return {
            'total_outstanding': float(total_outstanding),
            'total_unpaid_items': sum(item['unpaid_items'] for item in by_group),
            'by_group': by_group,
            'debtors': debtors,
            'top_debtor': debtors[0] if debtors else None,
        }

    def _build_unanswered_over_24h(self, director_user):
        threshold = timezone.now() - timedelta(hours=24)

        message_qs = Message.objects.filter(
            Q(sender=director_user) | Q(receiver=director_user)
        ).annotate(
            participant_id=Case(
                When(sender=director_user, then=F('receiver_id')),
                default=F('sender_id'),
                output_field=IntegerField(),
            )
        ).exclude(participant_id=director_user.id).select_related('sender', 'receiver').order_by('participant_id', '-created_at')

        latest_by_participant = {}
        for message in message_qs:
            participant_id = int(message.participant_id)
            if participant_id not in latest_by_participant:
                latest_by_participant[participant_id] = message

        pending = []
        now = timezone.now()
        for participant_id, last_message in latest_by_participant.items():
            if last_message.sender_id == director_user.id:
                continue
            if last_message.created_at > threshold:
                continue

            participant = last_message.sender
            if participant.id == director_user.id:
                participant = last_message.receiver

            if not participant or not participant.is_parent:
                continue

            full_name = f"{participant.first_name} {participant.last_name}".strip() or participant.username
            hours_waiting = int((now - last_message.created_at).total_seconds() // 3600)
            pending.append({
                'participant_id': int(participant.id),
                'participant_name': full_name,
                'last_message_preview': (last_message.body or '')[:100],
                'last_message_at': last_message.created_at.isoformat(),
                'hours_waiting': max(hours_waiting, 24),
            })

        pending.sort(key=lambda item: item['last_message_at'])
        return pending

    def get(self, request):
        today = timezone.localdate()
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=4)

        # 1. Liczba nieprzeczytanych wiadomości (skierowanych do dyrekcji)
        unread_messages_count = Message.objects.filter(
            receiver=request.user,
            is_read=False
        ).count()

        # 2. Liczba zgłoszonych nieobecności w bieżącym tygodniu roboczym (poniedziałek-piątek)
        absent_week_count = Attendance.objects.filter(
            date__range=(week_start, week_end),
            status='absent'
        ).count()

        # 3. Całkowita liczba dzieci
        total_children_count = Child.objects.count()
        
        # 4. Liczba obecnych (Total - Nieobecni)
        present_today_count = total_children_count - Attendance.objects.filter(
            date=today,
            status='absent'
        ).count()

        # 5. Frekwencja tygodniowa/miesięczna (cała placówka + każda grupa)
        groups = list(Group.objects.order_by('name').values('id', 'name'))
        attendance_week = {'all': self._build_attendance_series(7)}
        attendance_month = {'all': self._build_attendance_series(30)}

        for group in groups:
            group_id = int(group['id'])
            key = str(group_id)
            attendance_week[key] = self._build_attendance_series(7, group_id=group_id)
            attendance_month[key] = self._build_attendance_series(30, group_id=group_id)

        # 6. Zaległości i najwięksi dłużnicy
        debt_stats = self._build_debt_stats()

        # 7. Rozmowy bez odpowiedzi >24h
        unanswered = self._build_unanswered_over_24h(request.user)
        
        # Przygotowujemy dane do wysłania
        stats = {
            'unread_messages': unread_messages_count,
            'absent_today': absent_week_count,
            'absent_week': absent_week_count,
            'present_today': present_today_count,
            'total_children': total_children_count,
            'attendance': {
                'groups': [{'id': 'all', 'name': 'Cała placówka'}] + [
                    {'id': str(group['id']), 'name': group['name']}
                    for group in groups
                ],
                'week': attendance_week,
                'month': attendance_month,
            },
            'debts': debt_stats,
            'unanswered_over_24h': unanswered,
            'unanswered_over_24h_count': len(unanswered),
        }
        
        return Response(stats)

from rest_framework.permissions import IsAuthenticated
from .models import Preschool
from .serializers import PreschoolSerializer

class PreschoolViewSet(viewsets.ModelViewSet):
    queryset = Preschool.objects.all()
    serializer_class = PreschoolSerializer
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'put', 'patch']
    # Odczyt i edycja (brak tworzenia/kasowania przez API)